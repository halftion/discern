# coding=utf-8
import nltk
import os
import openpyxl
import re
import json

tempjson = './temp.json'
dir = r'C:\Users\AppData\Roaming\nltk_data'
nltk.data.path.append(dir)

def zmapScan(iplist):
    result = {}
    try:
        os.system("zgrab2 multiple -c multiple.ini")
        with open('./iplist.txt', 'r', encoding='utf-8') as file:
            json.dump(iplist, file, ident=4)
        os.system("zmap -P 80 -r 1000 -o active.json iplist.txt")
        os.system("zgrab2 --input-file=active.json --output-file=result.json --sonders=1000 http")
        with open('./result.json', 'r', encoding='utf-8') as file:
            result = json.load(file)
    except Exception as e:
        print(e)
    return result



# 结果处理
def loadresults(results):
    templist = []
    # print(len(temp))
    for t in results:
        data = str(t['data'])
        # print(data)
        # 删除不可打印字符
        data = re.sub(r'\\[n|r|t|v|f|s|S|cx]', '', data)
        # 删除http标签
        data = re.sub(r'<[^<]+?>', '', data)
            # 删除标点符号
        data = data.replace('@', '')
        data = data.replace('\\"', '@')
        data = re.sub(r'[\s+\!\\\/=|_@$&#%^*(+\')]+', '', data)
        data = data.replace("\"", "$").replace(",", "%").replace('[', '#').replace(']', '&')
        # print(data)
        # 分词
        word = nltk.word_tokenize(data)
        # 删除分隔符
        cutword = ["$", "%", '#', '&', '{', '}', ':']
        word = [w for w in word if not w in cutword]
        # 删除停止词
        stop_words = set(nltk.corpus.stopwords.words('english'))
        stop_words.remove(u'will')
        stop_words.remove(u'do')
        filtered_sentence = [w for w in word if not w in stop_words]
        data = filtered_sentence
        # print(data)
        templist.append({'ip':t['ip'],'data':data})
        # print(templist)
        return templist

def getproduct(db):
    return db

# 读取指纹数据库表单:
def loadxls(filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook[workbook.sheetnames[0]]
    # 建立指纹数据字典
    temp = {}
    for i in range(2,worksheet.max_row + 1):
        brand = str(worksheet['A' + str(i)].value)
        list = worksheet['B' + str(i)].value.replace('\'', '').split(', ')
        temp[brand] = list
        print("\rloadlig:" + str(i-1) + "/" + str(worksheet.max_row),end = '')
    print("\rload xls successful")
    return temp

# 关键词分析
def analywordlist(keyword,wordlist):
    for word in wordlist:
        if keyword in word:
            return True
    return False
# 设备关键词匹配
def analysis(db,json):
    fingerprint = getproduct(db)
    results = loadresults(json)
    devicelist = []
    for result in results:
        typeModel = ''
        deviceBrand = ''
        deviceType = ''
        for type in fingerprint.keys():
            if analywordlist(type,result['data']):
                deviceType = type
                for brand in fingerprint[deviceType]:
                    # print("brand:" + brand)
                    if analywordlist(brand, result['data']):
                        deviceBrand = brand
                        for model in fingerprint[brand]:
                            # print('model' + model)
                            if analywordlist(model, result['data']):
                                typeModel = model
                                break
                        break
        device = {'ip':result['ip'],'type':deviceType,'brand':deviceBrand,'model':typeModel}
        devicelist.append(device)
    return devicelist


# if __name__ == '__main__':
#